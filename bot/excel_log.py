"""Mantiene un file Excel (trades.xlsx) che si arricchisce di una riga
ogni volta che il bot completa un trade, più un foglio riassuntivo con
lo stato del portafoglio aggiornato ad ogni esecuzione (anche senza
trade), utile per vedere l'andamento nel tempo con un grafico in Excel.
"""
import os
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

TRADES_SHEET = "Trades"
SNAPSHOT_SHEET = "Andamento"

TRADE_HEADERS = [
    "Data/Ora (UTC)", "Tipo", "Prezzo mercato", "Prezzo eseguito",
    "Importo USD", "Quantità ETH", "Commissione USD", "Motivo",
    "Cash dopo", "ETH dopo", "Equity dopo (USD)", "Return cumulato %"
]

SNAPSHOT_HEADERS = [
    "Data/Ora (UTC)", "Prezzo ETH", "Cash (USD)", "ETH detenuto",
    "Equity totale (USD)", "Return cumulato %", "Drawdown attuale %",
    "In posizione", "N. piramidi", "Trade totali"
]


def _init_workbook(filepath):
    wb = Workbook()
    ws_trades = wb.active
    ws_trades.title = TRADES_SHEET
    ws_trades.append(TRADE_HEADERS)
    for col_idx, header in enumerate(TRADE_HEADERS, 1):
        cell = ws_trades.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    ws_snap = wb.create_sheet(SNAPSHOT_SHEET)
    ws_snap.append(SNAPSHOT_HEADERS)
    for col_idx, header in enumerate(SNAPSHOT_HEADERS, 1):
        cell = ws_snap.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    wb.save(filepath)
    return wb


def _autofit_columns(ws):
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def append_trade(filepath, side, market_price, exec_price, notional_usd,
                  amount_eth, fee_usd, reason, cash_after, eth_after,
                  equity_after, initial_capital):
    if not os.path.exists(filepath):
        wb = _init_workbook(filepath)
    else:
        wb = load_workbook(filepath)

    ws = wb[TRADES_SHEET]
    ret_pct = (equity_after / initial_capital - 1) * 100
    ws.append([
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        side, round(market_price, 2), round(exec_price, 2),
        round(notional_usd, 2), round(amount_eth, 6), round(fee_usd, 4),
        reason, round(cash_after, 2), round(eth_after, 6),
        round(equity_after, 2), round(ret_pct, 2)
    ])
    _autofit_columns(ws)
    wb.save(filepath)


def append_snapshot(filepath, price, cash, eth, equity, initial_capital,
                     drawdown, in_position, n_pyramids, trades_executed):
    if not os.path.exists(filepath):
        wb = _init_workbook(filepath)
    else:
        wb = load_workbook(filepath)

    ws = wb[SNAPSHOT_SHEET]
    ret_pct = (equity / initial_capital - 1) * 100
    ws.append([
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        round(price, 2), round(cash, 2), round(eth, 6), round(equity, 2),
        round(ret_pct, 2), round(drawdown * 100, 2),
        "SI" if in_position else "NO", n_pyramids, trades_executed
    ])
    _autofit_columns(ws)
    wb.save(filepath)
